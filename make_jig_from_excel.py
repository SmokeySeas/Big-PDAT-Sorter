# make_jig_from_excel.py
# Reads the master stud list Excel, extracts studs for a given Jig #,
# transforms global coords → local 2D coords, generates jig_data.scad,
# and exports STL via OpenSCAD CLI.

from __future__ import annotations

import math
import subprocess
from pathlib import Path

import openpyxl

# --- Config ---
OPENS = "openscad"
EXCEL_PATH = Path(r"C:\Users\broy7\Downloads\Master list Stud List Phase 3 - updated as of 01_29_2025 DS.xlsx")
SHEET_NAME = "Millenium Phase 3"
TARGET_JIG = 1

# Stud type → clearance hole diameter (mm)
# These are clearance holes, not the stud OD itself
STUD_CLEARANCE = {
    "T5x14.2": 6.0,       # T5 stud → ~5mm OD, 6mm clearance
    "M6x16": 7.0,         # M6 stud → 6mm OD, 7mm clearance
    "M6-Massebolzen": 7.0, # M6 ground bolt
    "M8x15": 9.0,         # M8 stud → 8mm OD, 9mm clearance
}

# Plate parameters
PLATE_THICKNESS = 8.0   # mm
PAD_DIAMETER = 16.0     # mm
PAD_HEIGHT = 6.0        # mm
MARGIN = 20.0           # mm border around stud pattern


def load_jig_studs(excel_path: Path, sheet: str, jig_num: int):
    """Load studs for a specific jig number from the Excel."""
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb[sheet]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    idx_id = headers.index("Stud ID#")
    idx_x = headers.index("PT_x")
    idx_y = headers.index("PT_y")
    idx_z = headers.index("PT_z")
    idx_type = headers.index("Stud Type")
    idx_jig = headers.index("Jig #")

    studs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx_jig] == jig_num:
            studs.append({
                "id": row[idx_id],
                "gx": float(row[idx_x]),
                "gy": float(row[idx_y]),
                "gz": float(row[idx_z]),
                "type": str(row[idx_type]),
            })
    wb.close()
    return studs


def global_to_local_2d(studs: list[dict]) -> list[dict]:
    """
    Convert global 3D coordinates to local 2D jig coordinates.

    Strategy for firewall-type surfaces:
    - The studs share roughly the same X (depth into vehicle)
    - Y and Z vary along the surface
    - Local X = global Y direction (lateral)
    - Local Y = global Z direction (vertical)
    - Origin = first stud position (shifted so all coords are positive)
    """
    if len(studs) < 1:
        return []

    # Use Y and Z as the 2D plane (X is roughly constant = surface normal)
    ys = [s["gy"] for s in studs]
    zs = [s["gz"] for s in studs]

    # Shift origin so minimum is at margin
    min_y = min(ys)
    min_z = min(zs)

    local_studs = []
    for s in studs:
        local_studs.append({
            "id": s["id"],
            "local_x": s["gy"] - min_y + MARGIN,
            "local_y": s["gz"] - min_z + MARGIN,
            "diameter": STUD_CLEARANCE.get(s["type"], 7.0),
            "type": s["type"],
        })

    return local_studs


def compute_plate_size(local_studs):
    """Compute plate dimensions from stud bounding box + margin."""
    xs = [s["local_x"] for s in local_studs]
    ys = [s["local_y"] for s in local_studs]
    plate_x = max(xs) - min(xs) + 2 * MARGIN
    plate_y = max(ys) - min(ys) + 2 * MARGIN
    return plate_x, plate_y


def place_pads(plate_x, plate_y):
    """Place 3 pads in a stable triangle pattern."""
    return [
        (MARGIN, MARGIN, PAD_DIAMETER, PAD_HEIGHT),
        (plate_x - MARGIN, MARGIN, PAD_DIAMETER, PAD_HEIGHT),
        (plate_x / 2.0, plate_y - MARGIN, PAD_DIAMETER, PAD_HEIGHT),
    ]


def write_jig_data_scad(out_path: Path, plate, studs, pads, jig_num, stud_info):
    """Write the generated .scad data file."""
    def fmt3(arr):
        return "[" + ",".join(f"[{a:.3f},{b:.3f},{c:.3f}]" for a, b, c in arr) + "]"

    def fmt4(arr):
        return "[" + ",".join(f"[{a:.3f},{b:.3f},{c:.3f},{d:.3f}]" for a, b, c, d in arr) + "]"

    stud_tuples = [(s["local_x"], s["local_y"], s["diameter"]) for s in studs]

    lines = [
        f"// Auto-generated for Jig #{jig_num}",
    ]
    for s in stud_info:
        lines.append(f"// Stud {s['id']}: {s['type']} @ global ({s['gx']}, {s['gy']}, {s['gz']})")
    lines.append("")
    lines.append(f"plate = [{plate[0]:.3f},{plate[1]:.3f},{plate[2]:.3f}];")
    lines.append(f"studs = {fmt3(stud_tuples)};")
    lines.append(f"pads  = {fmt4(pads)};")
    lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


def export_stl(base_scad: Path, stl_out: Path):
    cmd = [OPENS, "-o", str(stl_out), str(base_scad)]
    print(f"Running: {' '.join(cmd)}")
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"STDERR: {result.stderr}")
        raise RuntimeError(f"OpenSCAD failed with code {result.returncode}")


if __name__ == "__main__":
    root = Path(__file__).parent
    gen_dir = root / "generated"
    out_dir = root / "out"
    gen_dir.mkdir(exist_ok=True)
    out_dir.mkdir(exist_ok=True)

    # Step 1: Load studs from Excel
    print(f"Loading Jig #{TARGET_JIG} from {EXCEL_PATH.name}...")
    raw_studs = load_jig_studs(EXCEL_PATH, SHEET_NAME, TARGET_JIG)
    print(f"Found {len(raw_studs)} studs:")
    for s in raw_studs:
        print(f"  {s['id']}: ({s['gx']}, {s['gy']}, {s['gz']}) {s['type']}")

    # Step 2: Global → Local transform
    local_studs = global_to_local_2d(raw_studs)
    print(f"\nLocal coordinates:")
    for s in local_studs:
        print(f"  {s['id']}: ({s['local_x']:.1f}, {s['local_y']:.1f}) hole={s['diameter']}mm")

    # Step 3: Compute plate + pads
    plate_x, plate_y = compute_plate_size(local_studs)
    plate = (plate_x, plate_y, PLATE_THICKNESS)
    pads = place_pads(plate_x, plate_y)
    print(f"\nPlate: {plate_x:.1f} x {plate_y:.1f} x {PLATE_THICKNESS} mm")

    # Step 4: Write generated .scad
    data_path = gen_dir / "jig_data.scad"
    write_jig_data_scad(data_path, plate, local_studs, pads, TARGET_JIG, raw_studs)
    print(f"Wrote: {data_path}")

    # Step 5: Export STL
    stl_name = f"jig_{TARGET_JIG:03d}.stl"
    stl_out = out_dir / stl_name
    export_stl(root / "base_jig_pads.scad", stl_out)
    print(f"\nDone! Open: {stl_out}")
